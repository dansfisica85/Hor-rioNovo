import openpyxl, random, math

SRC = "Manha_Ferrucio_2026_SEM_CHOQUES.xlsx"
wb = openpyxl.load_workbook(SRC)
ws = wb.active
NR, NC = ws.max_row, ws.max_column

LOCKED = {"DRIELLI","ANDREA","ALEXANDRE","SAMIA","ROBERTA",
          "LILIANE","DINO","MARIANA","SELMA"}
EXEMPT = {"ALINE"}

def val(r,c):
    x=ws.cell(r,c).value
    return "" if x is None else str(x).strip()
def setv(r,c,s):
    ws.cell(r,c).value = s if s!="" else None
def teacher(s):
    if not s or "RECREIO" in s.upper(): return None
    if "-" in s: return s.split("-")[0].strip().upper()
    return None
def standalone(s):
    return bool(s) and "-" not in s and "RECREIO" not in s.upper()

TROWS=[r for r in range(2,NR+1) if "RECREIO" not in val(r,2).upper()]

order=[]
for r in range(2,NR+1):
    d=val(r,1)
    if d: order.append((d,[]))
    order[-1][1].append(None if "RECREIO" in val(r,2).upper() else r)

def movable_cell(r,c):
    s=val(r,c)
    if not s or standalone(s): return False
    t=teacher(s)
    if t in LOCKED or t=="BAQUETE": return False
    return True

def row_conflict(r):
    seen=set()
    for c in range(3,NC+1):
        t=teacher(val(r,c))
        if not t or t in EXEMPT: continue
        if t in seen: return True
        seen.add(t)
    return False

def baquete_friday():
    for d,rows in order:
        if d.upper()!="SEXTA": continue
        for r in rows:
            if r is None: continue
            for c in range(3,NC+1):
                if teacher(val(r,c))=="BAQUETE": return True
    return False

def cost():
    total=0
    for d,rows in order:
        present=[]
        for r in rows:
            if r is None: present.append(None); continue
            s=set()
            for c in range(3,NC+1):
                t=teacher(val(r,c))
                if t and t not in EXEMPT and t not in LOCKED: s.add(t)
            present.append(s)
        teachers=set()
        for s in present:
            if s: teachers|=s
        for t in teachers:
            run=0
            for s in present:
                if s is None:
                    if run>=4: total+=(run-2)*100
                    elif run==3: total+=8
                    run=0; continue
                if t in s: run+=1
                else:
                    if run>=4: total+=(run-2)*100
                    elif run==3: total+=8
                    run=0
            if run>=4: total+=(run-2)*100
            elif run==3: total+=8
    return total

# movable cells grouped by column
colcells={c:[r for r in TROWS if movable_cell(r,c)] for c in range(3,NC+1)}
cols=[c for c in colcells if len(colcells[c])>=2]

def random_swap():
    c=random.choice(cols)
    r1,r2=random.sample(colcells[c],2)
    if val(r1,c)==val(r2,c): return None
    a,b=val(r1,c),val(r2,c)
    setv(r1,c,b); setv(r2,c,a)
    if row_conflict(r1) or row_conflict(r2):
        setv(r1,c,a); setv(r2,c,b); return None
    # baquete cannot land on friday: r1/r2 are non-baquete cells, but partner could be baquete? no, both movable (not baquete). safe.
    return (r1,r2,c,a,b)

def undo(s):
    r1,r2,c,a,b=s
    setv(r1,c,a); setv(r2,c,b)

random.seed(7)
cur=cost(); best=cur
def snapshot():
    return {(r,c):val(r,c) for r in TROWS for c in range(3,NC+1)}
def restore(snap):
    for (r,c),v in snap.items(): setv(r,c,v)
best_snap=snapshot()

total_it=0
for restart in range(12):
    # reheat; start each restart from global best
    restore(best_snap); cur=best
    T=30.0; Tmin=0.02; alpha=0.99985
    while T>Tmin:
        total_it+=1
        s=random_swap()
        if s is None:
            T*=alpha; continue
        nc=cost()
        delta=nc-cur
        if delta<=0 or random.random()<math.exp(-delta/T):
            cur=nc
            if nc<best:
                best=nc; best_snap=snapshot()
        else:
            undo(s)
        T*=alpha
    print(f"  restart {restart}: best={best}")
restore(best_snap)
print(f"iters={total_it}  custo_final={best}")
assert not baquete_friday(), "BAQUETE caiu na sexta!"
# verify no conflicts
bad=[r for r in TROWS if row_conflict(r)]
print("linhas com choque (excl ALINE):", bad if bad else "NENHUMA ✅")
wb.save(SRC)
print("Salvo.")
