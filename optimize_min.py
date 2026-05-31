import openpyxl

SRC = "Manha_Ferrucio_2026_SEM_CHOQUES.xlsx"
wb = openpyxl.load_workbook(SRC)
ws = wb.active
NR, NC = ws.max_row, ws.max_column

LOCKED = {"DRIELLI","ANDREA","ALEXANDRE","SAMIA","ROBERTA",
          "LILIANE","DINO","MARIANA","SELMA"}
EXEMPT = {"ALINE"}

def val(r,c):
    x=ws.cell(r,c).value
    return "" if x is None else str(x).strip()
def setv(r,c,s):
    ws.cell(r,c).value = s if s!="" else None
def teacher(s):
    if not s or "RECREIO" in s.upper(): return None
    if "-" in s: return s.split("-")[0].strip().upper()
    return None
def standalone(s):
    return bool(s) and "-" not in s and "RECREIO" not in s.upper()

TROWS=[r for r in range(2,NR+1) if "RECREIO" not in val(r,2).upper()]
order=[]
for r in range(2,NR+1):
    d=val(r,1)
    if d: order.append((d,[]))
    order[-1][1].append(None if "RECREIO" in val(r,2).upper() else r)

def movable(r,c):
    s=val(r,c)
    if not s or standalone(s): return False
    t=teacher(s)
    if t in LOCKED or t=="BAQUETE": return False
    return True

def row_conflict(r):
    seen=set()
    for c in range(3,NC+1):
        t=teacher(val(r,c))
        if not t or t in EXEMPT: continue
        if t in seen: return True
        seen.add(t)
    return False

def baquete_friday():
    for d,rows in order:
        if d.upper()!="SEXTA": continue
        for r in rows:
            if r is None: continue
            for c in range(3,NC+1):
                if teacher(val(r,c))=="BAQUETE": return True
    return False

def cost():
    total=0
    for d,rows in order:
        present=[]
        for r in rows:
            if r is None: present.append(None); continue
            s=set()
            for c in range(3,NC+1):
                t=teacher(val(r,c))
                if t and t not in EXEMPT and t not in LOCKED: s.add(t)
            present.append(s)
        teachers=set()
        for s in present:
            if s: teachers|=s
        for t in teachers:
            run=0
            for s in present:
                if s is None:
                    if run>=4: total+=(run-2)*1000
                    elif run==3: total+=10
                    run=0; continue
                if t in s: run+=1
                else:
                    if run>=4: total+=(run-2)*1000
                    elif run==3: total+=10
                    run=0
            if run>=4: total+=(run-2)*1000
            elif run==3: total+=10
    return total

# Best-improvement hill climb: only accept STRICT improvements (minimal churn)
def best_improve():
    cur=cost()
    improved=True; nmoves=0
    while improved:
        improved=False
        best_gain=0; best=None
        for c in range(3,NC+1):
            rows=[r for r in TROWS if movable(r,c)]
            for i in range(len(rows)):
                for j in range(i+1,len(rows)):
                    r1,r2=rows[i],rows[j]
                    a,b=val(r1,c),val(r2,c)
                    if a==b: continue
                    setv(r1,c,b); setv(r2,c,a)
                    if row_conflict(r1) or row_conflict(r2) or baquete_friday():
                        setv(r1,c,a); setv(r2,c,b); continue
                    nc=cost()
                    setv(r1,c,a); setv(r2,c,b)
                    gain=cur-nc
                    if gain>best_gain:
                        best_gain=gain; best=(r1,r2,c,a,b)
        if best:
            r1,r2,c,a,b=best
            setv(r1,c,b); setv(r2,c,a)
            cur-=best_gain; nmoves+=1; improved=True
    return cur,nmoves

print("Custo inicial:", cost())
final,nm=best_improve()
print(f"Custo final: {final}  (trocas aplicadas: {nm})")
assert not baquete_friday()
print("Choques:", "NENHUM" if not any(row_conflict(r) for r in TROWS) else "ERRO")
wb.save(SRC)
print("Salvo.")
