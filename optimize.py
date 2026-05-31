import openpyxl, glob, random

SRC = "Manha_Ferrucio_2026_SEM_CHOQUES.xlsx"
wb = openpyxl.load_workbook(SRC)
ws = wb.active
NR, NC = ws.max_row, ws.max_column

LOCKED = {"DRIELLI","ANDREA","ALEXANDRE","SAMIA","ROBERTA",
          "LILIANE","DINO","MARIANA","SELMA"}
EXEMPT = {"ALINE"}  # ignored for consecutive-rule & conflict

def val(r,c):
    x=ws.cell(r,c).value
    return "" if x is None else str(x).strip()
def setv(r,c,s):
    ws.cell(r,c).value = s if s!="" else None
def teacher(s):
    if not s or "RECREIO" in s.upper(): return None
    if "-" in s: return s.split("-")[0].strip().upper()
    return None
def standalone(s):
    return bool(s) and "-" not in s and "RECREIO" not in s.upper()

TROWS = [r for r in range(2,NR+1) if "RECREIO" not in val(r,2).upper()]
def day_of(r):
    rr=r
    while rr>=2 and not val(rr,1): rr-=1
    return val(rr,1).upper()

# ordered day structure with break markers (for consecutive runs)
DAYSEQ={}  # day -> list of (rowOrNone) in order, None = recreio break
order=[]
for r in range(2,NR+1):
    d=val(r,1)
    if d: order.append((d,[]))
    if "RECREIO" in val(r,2).upper():
        order[-1][1].append(None)
    else:
        order[-1][1].append(r)

def movable_cell(r,c):
    s=val(r,c)
    if not s: return False
    if standalone(s): return False
    t=teacher(s)
    if t in LOCKED: return False
    if t=="BAQUETE": return False
    return True

def row_conflict(r):
    seen={}
    for c in range(3,NC+1):
        t=teacher(val(r,c))
        if not t or t in EXEMPT: continue
        seen.setdefault(t,0)
        seen[t]+=1
    return any(n>1 for n in seen.values())

def baquete_on_friday():
    for d,rows in order:
        if d.upper()!="SEXTA": continue
        for r in rows:
            if r is None: continue
            for c in range(3,NC+1):
                if teacher(val(r,c))=="BAQUETE": return True
    return False

def cost():
    """Penalize long consecutive runs for movable (non-locked,non-exempt) teachers."""
    total=0
    # presence per (day,slotindex)
    for d,rows in order:
        # gather teacher presence per slot
        present=[]
        for r in rows:
            if r is None:
                present.append(None); continue
            s=set()
            for c in range(3,NC+1):
                t=teacher(val(r,c))
                if t and t not in EXEMPT and t not in LOCKED:
                    s.add(t)
            present.append(s)
        teachers=set()
        for s in present:
            if s: teachers|=s
        for t in teachers:
            run=0
            for s in present:
                if s is None:
                    if run>=4: total += (run-2)*100
                    elif run==3: total += 5
                    run=0; continue
                if t in s: run+=1
                else:
                    if run>=4: total += (run-2)*100
                    elif run==3: total += 5
                    run=0
            if run>=4: total += (run-2)*100
            elif run==3: total += 5
    return total

def try_swap(r1,r2,c):
    a,b=val(r1,c),val(r2,c)
    setv(r1,c,b); setv(r2,c,a)
    if row_conflict(r1) or row_conflict(r2) or baquete_on_friday():
        setv(r1,c,a); setv(r2,c,b); return False
    return True
def undo_swap(r1,r2,c):
    a,b=val(r1,c),val(r2,c)
    setv(r1,c,b); setv(r2,c,a)

print("Custo inicial:", cost())
random.seed(42)
best=cost()
improved=True
passes=0
while improved and passes<200:
    improved=False; passes+=1
    cols=list(range(3,NC+1)); random.shuffle(cols)
    for c in cols:
        rows=[r for r in TROWS if movable_cell(r,c)]
        random.shuffle(rows)
        for i in range(len(rows)):
            for j in range(i+1,len(rows)):
                r1,r2=rows[i],rows[j]
                if val(r1,c)==val(r2,c): continue
                if try_swap(r1,r2,c):
                    nc=cost()
                    if nc<best:
                        best=nc; improved=True
                    else:
                        undo_swap(r1,r2,c)
    # early exit if no 4+ runs remain (cost from 3-runs only, <100)
print(f"Custo final: {best} (passes={passes})")

wb.save(SRC)
print("Salvo.")
