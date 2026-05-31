import openpyxl, random

SRC = "Manha_Ferrucio_2026_SEM_CHOQUES.xlsx"
wb = openpyxl.load_workbook(SRC)
ws = wb.active
NR, NC = ws.max_row, ws.max_column
LOCKED={"DRIELLI","ANDREA","ALEXANDRE","SAMIA","ROBERTA","LILIANE","DINO","MARIANA","SELMA"}
EXEMPT={"ALINE"}
def val(r,c):
    x=ws.cell(r,c).value; return "" if x is None else str(x).strip()
def setv(r,c,s): ws.cell(r,c).value = s if s!="" else None
def teacher(s):
    if not s or "RECREIO" in s.upper(): return None
    return s.split("-")[0].strip().upper() if "-" in s else None
def standalone(s): return bool(s) and "-" not in s and "RECREIO" not in s.upper()
TROWS=[r for r in range(2,NR+1) if "RECREIO" not in val(r,2).upper()]
order=[]
for r in range(2,NR+1):
    d=val(r,1)
    if d: order.append((d,[]))
    order[-1][1].append(None if "RECREIO" in val(r,2).upper() else r)
def movable(r,c):
    s=val(r,c)
    if not s or standalone(s): return False
    t=teacher(s)
    return not (t in LOCKED or t=="BAQUETE")
def row_conflict(r):
    seen=set()
    for c in range(3,NC+1):
        t=teacher(val(r,c))
        if not t or t in EXEMPT: continue
        if t in seen: return True
        seen.add(t)
    return False
def baquete_friday():
    for d,rows in order:
        if d.upper()!="SEXTA": continue
        for r in rows:
            if r is None: continue
            for c in range(3,NC+1):
                if teacher(val(r,c))=="BAQUETE": return True
    return False
def cost():
    total=0
    for d,rows in order:
        present=[]
        for r in rows:
            if r is None: present.append(None); continue
            s=set()
            for c in range(3,NC+1):
                t=teacher(val(r,c))
                if t and t not in EXEMPT and t not in LOCKED: s.add(t)
            present.append(s)
        ts=set()
        for s in present:
            if s: ts|=s
        for t in ts:
            run=0
            for s in present:
                if s is None:
                    if run>=4: total+=(run-2)*1000
                    elif run==3: total+=10
                    run=0; continue
                if t in s: run+=1
                else:
                    if run>=4: total+=(run-2)*1000
                    elif run==3: total+=10
                    run=0
            if run>=4: total+=(run-2)*1000
            elif run==3: total+=10
    return total
COLS=[c for c in range(3,NC+1)]
def valid_swap(r1,r2,c):
    a,b=val(r1,c),val(r2,c)
    if a==b: return False
    setv(r1,c,b); setv(r2,c,a)
    if row_conflict(r1) or row_conflict(r2) or baquete_friday():
        setv(r1,c,a); setv(r2,c,b); return False
    return True
def best_improve():
    cur=cost(); improved=True
    while improved:
        improved=False; bg=0; best=None
        for c in COLS:
            rows=[r for r in TROWS if movable(r,c)]
            for i in range(len(rows)):
                for j in range(i+1,len(rows)):
                    r1,r2=rows[i],rows[j]
                    if not valid_swap(r1,r2,c): continue
                    nc=cost(); setv(r1,c,val(r2,c)); setv(r2,c,val(r1,c))  # undo via reswap
                    g=cur-nc
                    if g>bg: bg=g; best=(r1,r2,c)
        if best:
            r1,r2,c=best; a,b=val(r1,c),val(r2,c); setv(r1,c,b); setv(r2,c,a)
            cur-=bg; improved=True
    return cur
def snapshot(): return {(r,c):val(r,c) for r in TROWS for c in COLS}
def restore(sn):
    for (r,c),v in sn.items(): setv(r,c,v)

random.seed(123)
cur=best_improve(); best=cur; bsnap=snapshot()
print("start:",best)
for k in range(60):
    restore(bsnap)
    # perturb: 2-4 random neutral-ish swaps
    for _ in range(random.randint(2,4)):
        c=random.choice(COLS)
        rows=[r for r in TROWS if movable(r,c)]
        if len(rows)<2: continue
        r1,r2=random.sample(rows,2)
        valid_swap(r1,r2,c)  # apply if valid (may raise cost)
    c2=best_improve()
    if c2<best:
        best=c2; bsnap=snapshot(); print(f"  iter {k}: melhorou -> {best}")
restore(bsnap)
print("final:",best)
assert not baquete_friday()
assert not any(row_conflict(r) for r in TROWS)
wb.save(SRC)
print("Salvo.")
