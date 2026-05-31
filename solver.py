import openpyxl, glob, copy

SRC = glob.glob("Manha*.xlsx")[0]
OUT = "Manha_Ferrucio_2026_SEM_CHOQUES.xlsx"

wb = openpyxl.load_workbook(SRC)  # keep styles
ws = wb.active
NR, NC = ws.max_row, ws.max_column
headers = [ws.cell(1, c).value for c in range(1, NC + 1)]

# Locked teachers (cannot move at all)
LOCKED_TEACHERS = {
    "DRIELLI", "ANDREA", "ALEXANDRE", "SAMIA", "ROBERTA",
    "LILIANE", "DINO", "MARIANA", "SELMA",
}
# BAQUETE is locked EXCEPT must leave Friday
# ALINE is fully exempt from conflict rules and may be moved freely

def val(r, c):
    v = ws.cell(r, c).value
    return "" if v is None else str(v).strip()

def teacher(s):
    if not s:
        return None
    if "RECREIO" in s.upper():
        return None
    if "-" in s:
        return s.split("-")[0].strip().upper()
    return None  # standalone discipline

def is_standalone(s):
    return bool(s) and "-" not in s and "RECREIO" not in s.upper()

# timeslot rows (exclude header + recreio)
TROWS = [r for r in range(2, NR + 1) if "RECREIO" not in val(r, 2).upper()]

# day per row
def day_of(r):
    rr = r
    while rr >= 2 and not val(rr, 1):
        rr -= 1
    return val(rr, 1).upper()

# grid model: dict[(r,c)] = string ; we mutate ws directly via swaps
def set_cell(r, c, s):
    ws.cell(r, c).value = s if s != "" else None

def movable(r, c):
    """Cell content can be relocated (used as a swap partner)."""
    s = val(r, c)
    if not s:
        return False
    if is_standalone(s):
        return False  # locked discipline
    t = teacher(s)
    if t in LOCKED_TEACHERS:
        return False
    if t == "BAQUETE":
        return False  # don't shuffle BAQUETE as a partner
    return True

def conflicts_in_row(r, ignore=("ALINE",)):
    seen = {}
    for c in range(3, NC + 1):
        t = teacher(val(r, c))
        if not t or t in ignore:
            continue
        seen.setdefault(t, []).append(c)
    return {t: cs for t, cs in seen.items() if len(cs) > 1}

def teacher_in_row(t, r, except_col=None):
    for c in range(3, NC + 1):
        if c == except_col:
            continue
        if teacher(val(r, c)) == t:
            return True
    return False

def all_conflicts():
    out = {}
    for r in TROWS:
        d = conflicts_in_row(r)
        if d:
            out[r] = d
    return out

print("CONFLITOS INICIAIS (excl. ALINE):")
for r, d in all_conflicts().items():
    print(f"  R{r} {day_of(r)} {val(r,2)[:5]}: " +
          ", ".join(f"{t}->{[headers[c-1] for c in cs]}" for t, cs in d.items()))

# ---- Resolve DAVI conflicts via in-column swaps ----
def try_resolve_row(r, who):
    """Resolve a conflict of `who` in row r by relocating ONE of its cells."""
    cs = [c for c in range(3, NC + 1) if teacher(val(r, c)) == who]
    if len(cs) < 2:
        return True
    for c in cs:  # try moving each occurrence
        for r2 in TROWS:
            if r2 == r:
                continue
            if not movable(r2, c):
                continue
            # who must not already be in r2 (else still conflict after move)
            if teacher_in_row(who, r2):
                continue
            T = teacher(val(r2, c))  # displaced teacher
            # T moves to row r; must not already be in r (excluding the cell we vacate)
            if T != "ALINE" and teacher_in_row(T, r, except_col=c):
                continue
            # perform swap
            a, b = val(r, c), val(r2, c)
            set_cell(r, c, b)
            set_cell(r2, c, a)
            # verify neither row has NEW non-exempt conflict
            if not conflicts_in_row(r) and not conflicts_in_row(r2):
                return True
            # rollback
            set_cell(r, c, a)
            set_cell(r2, c, b)
    return False

davi_rows = [r for r in TROWS if "DAVI" in conflicts_in_row(r)]
for r in davi_rows:
    ok = try_resolve_row(r, "DAVI")
    print(f"DAVI R{r} ({day_of(r)} {val(r,2)[:5]}): {'RESOLVIDO' if ok else 'FALHOU'}")

# ---- Move BAQUETE off Friday ----
def baquete_friday_cells():
    out = []
    for r in TROWS:
        if day_of(r) != "SEXTA":
            continue
        for c in range(3, NC + 1):
            if teacher(val(r, c)) == "BAQUETE":
                out.append((r, c))
    return out

def move_baquete(r, c):
    for r2 in TROWS:
        if day_of(r2) == "SEXTA":
            continue
        if not movable(r2, c):
            continue
        if teacher_in_row("BAQUETE", r2):
            continue
        T = teacher(val(r2, c))
        # T goes to Friday row r (must be allowed & not already there)
        if T != "ALINE" and teacher_in_row(T, r, except_col=c):
            continue
        a, b = val(r, c), val(r2, c)
        set_cell(r, c, b)
        set_cell(r2, c, a)
        if not conflicts_in_row(r) and not conflicts_in_row(r2):
            return True
        set_cell(r, c, a)
        set_cell(r2, c, b)
    return False

for (r, c) in baquete_friday_cells():
    ok = move_baquete(r, c)
    print(f"BAQUETE saída sexta R{r} {headers[c-1]}: {'RESOLVIDO' if ok else 'FALHOU'}")

# ---- Final verification ----
print("\nCONFLITOS FINAIS (excl. ALINE):")
fin = all_conflicts()
if not fin:
    print("  NENHUM ✅")
else:
    for r, d in fin.items():
        print(f"  R{r}: {d}")
bf = baquete_friday_cells()
print("BAQUETE na sexta:", "NENHUM ✅" if not bf else bf)

wb.save(OUT)
print("\nSalvo em:", OUT)
