import openpyxl, glob, colorsys
from openpyxl.styles import PatternFill, Font

SRC = "Manha_Ferrucio_2026_SEM_CHOQUES.xlsx"
wb = openpyxl.load_workbook(SRC)
ws = wb.active
NR, NC = ws.max_row, ws.max_column

def val(r, c):
    v = ws.cell(r, c).value
    return "" if v is None else str(v).strip()

def key(s):
    """Return the legend key for a cell: teacher name, or discipline if standalone."""
    if not s or "RECREIO" in s.upper():
        return None
    if "-" in s:
        return s.split("-")[0].strip().upper()
    return s.strip().upper()  # standalone locked discipline

# collect all keys
keys = []
for r in range(2, NR + 1):
    for c in range(3, NC + 1):
        k = key(val(r, c))
        if k and k not in keys:
            keys.append(k)
keys.sort()
n = len(keys)
print(f"{n} entradas para colorir")

def color_for(i, n):
    h = (i / n) % 1.0
    # vary saturation/value slightly so adjacent hues differ more
    s = 0.45 + 0.20 * ((i * 3) % 5) / 4.0
    v = 0.97
    r, g, b = colorsys.hsv_to_rgb(h, s, v)
    return f"{int(r*255):02X}{int(g*255):02X}{int(b*255):02X}"

def luminance(hexc):
    r = int(hexc[0:2], 16); g = int(hexc[2:4], 16); b = int(hexc[4:6], 16)
    return (0.299 * r + 0.587 * g + 0.114 * b)

cmap = {k: color_for(i, n) for i, k in enumerate(keys)}

# apply fills
for r in range(2, NR + 1):
    for c in range(3, NC + 1):
        k = key(val(r, c))
        if not k:
            continue
        hexc = cmap[k]
        ws.cell(r, c).fill = PatternFill("solid", fgColor=hexc)
        font = ws.cell(r, c).font
        dark = luminance(hexc) > 150
        ws.cell(r, c).font = Font(name=font.name, size=font.size, bold=font.bold,
                                  color="000000" if dark else "FFFFFF")

# Build a legend on a new sheet
if "LEGENDA" in wb.sheetnames:
    del wb["LEGENDA"]
leg = wb.create_sheet("LEGENDA")
leg.cell(1, 1, "PROFESSOR / DISCIPLINA").font = Font(bold=True)
leg.cell(1, 2, "COR").font = Font(bold=True)
leg.column_dimensions["A"].width = 28
leg.column_dimensions["B"].width = 14
for i, k in enumerate(keys, start=2):
    hexc = cmap[k]
    leg.cell(i, 1, k)
    cell = leg.cell(i, 2, "")
    cell.fill = PatternFill("solid", fgColor=hexc)

wb.save(SRC)
print("Cores aplicadas e legenda criada em", SRC)
